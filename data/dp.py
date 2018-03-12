import openpyxl
from random import shuffle

infile_test = "test.enc"
outfile_test = "test.dec"
infile_train = "train.enc"
outfile_train = "train.dec"
readfile = "TalkOS.xlsx"
test = 20

if __name__ == '__main__':

    book = openpyxl.load_workbook(readfile)
    sheet = book.active
    rows = sheet.max_row
    test_size = test*(rows)/100

    indices = range(2, rows+1)
    for i in range(2000):
        shuffle(indices)
    test_set = indices[:test_size]
    train_set = indices[test_size:]
    
    print indices

    f = open(infile_test, 'w')
    g = open(outfile_test, 'w')

    for i in test_set:
        encoder = str(sheet.cell(row = i, column = 2).value) + "\n"
        decoder = str(sheet.cell(row = i, column = 3).value) + "\n"
        decoder = str(decoder)
        t = decoder.split(',')
        if len(t) > 1:
            decoder = t[0] + ' , ' + t[1]

        f.write(encoder)
        g.write(decoder)
    
    f.close()
    g.close()

    f = open(infile_train, 'w')
    g = open(outfile_train, 'w')

    for i in train_set:
        encoder = str(sheet.cell(row = i, column = 2).value) + "\n"
        decoder = str(sheet.cell(row = i, column = 3).value) + "\n"
        decoder = str(decoder)
        t = decoder.split(',')
        if len(t) > 1:
            decoder = t[0] + ' , ' + t[1]
        f.write(encoder)
        g.write(decoder)
    
    f.close()
    g.close()
